import time
import os
import tkinter as tk
from tkinter import ttk, messagebox

import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ===================== CONFIG =====================

GST_SEARCH_URL = "https://services.gst.gov.in/services/searchtp"
CHROMEDRIVER_PATH = r"C:\CromeDriver\chromedriver-win64\chromedriver.exe"

# CORRECTED FINAL LOCATORS
LOCATORS = {
    "gstin_input": (By.ID, "for_gstin"),

    "legal_name": (
        By.XPATH,
        '//strong[contains(text(),"Legal Name")]/parent::p/following-sibling::p[1]'
    ),
    "trade_name": (
        By.XPATH,
        '//strong[contains(text(),"Trade Name")]/parent::p/following-sibling::p[1]'
    ),
    "reg_status": (
        By.XPATH,
        '//strong[contains(text(),"GSTIN / UIN Status")]/parent::p/following-sibling::p[1]'
    ),

    # Buttons from your inspected HTML
    "show_filing_table_btn": (By.ID, "filingTable"),
    "filing_search_btn": (By.CSS_SELECTOR, "button.srchbtn"),

    # CORRECTED TABLE LOCATORS
    "gstr3b_info": (
    By.XPATH,
    '//h4[contains(text(),"GSTR3B")]/following::table[1]'
    ),

    "gstr1_info": (
        By.XPATH,
        '//h4[contains(text(),"GSTR-1") or contains(text(),"GSTR1") or contains(text(),"IFF")]/following::table[1]'
    ),
}


# ===================== DRIVER =====================

def create_driver():
    service = Service(CHROMEDRIVER_PATH)
    options = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=options)
    driver.maximize_window()
    return driver


def safe_get_text(driver, by, value):
    try:
        return driver.find_element(by, value).text.strip()
    except:
        return ""


# ===================== MAIN APP =====================

class GSTFilingApp:
     def __init__(self, root):
        self.root = root
        self.root.title("GST Filing Data Collector")
        self.root.geometry("700x450")

        self.driver = None
        self.create_ui()

     def create_ui(self):
        ttk.Label(self.root, text="GST Filing Data Collector",
                  font=("Segoe UI", 18, "bold")).pack(pady=10)

        ttk.Label(
            self.root,
            text="1. Enter GSTIN(s)\n"
                 "2. Click Fetch\n"
                 "3. Enter CAPTCHA manually & click the FIRST Search\n"
                 "4. After result loads, click OK\n"
                 "5. Program extracts latest GSTR-1 & GSTR-3B dates automatically",
            justify="left"
        ).pack(anchor="w", padx=10)

        frame = ttk.Frame(self.root)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        ttk.Label(frame, text="GSTINs (one per line):").pack(anchor="w")
        self.txt_gstin = tk.Text(frame, height=10)
        self.txt_gstin.pack(fill="both", expand=True)

        btns = ttk.Frame(self.root)
        btns.pack(pady=10)

        ttk.Button(btns, text="Fetch & Save", command=self.on_fetch).pack(side="left", padx=5)
        ttk.Button(btns, text="Exit", command=self.on_exit).pack(side="left")

        self.status = tk.StringVar(value="Ready.")
        ttk.Label(self.root, textvariable=self.status, relief="sunken").pack(fill="x", side="bottom")

     def update_status(self, t):
        self.status.set(t)
        self.root.update_idletasks()

    # ===================== SAVE EXCEL (Fixed File, Perfect Format) =====================

    
     def save_excel(self, rows):
        # Always save on Desktop
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        file_path = os.path.join(desktop, "gst_filing_data.xlsx")

        headers = ["GSTIN", "Legal Name", "Trade Name",
                   "Return Filing Status", "GSTR-1 Info",
                   "GSTR-3B Info", "Registration Status"]

        if not os.path.exists(file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "GST Data"

            for col, h in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = h
                ws.cell(row=1, column=col).font = Font(bold=True)

            wb.save(file_path)

        wb = load_workbook(file_path)
        ws = wb.active

        ws.delete_rows(2, ws.max_row)

        for r_idx, row_data in enumerate(rows, start=2):
            for c_idx, val in enumerate(row_data.values(), start=1):
                ws.cell(row=r_idx, column=c_idx).value = val

        # Auto width:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3

        wb.save(file_path)
        return file_path


    # ===================== FETCH CLICK =====================

     def on_fetch(self):
        gstins = [g.strip() for g in self.txt_gstin.get("1.0", "end").splitlines() if g.strip()]
        if not gstins:
            messagebox.showwarning("GSTIN missing", "Enter GSTIN(s).")
            return

        if self.driver is None:
            self.driver = create_driver()

        all_rows = []

        for i, gstin in enumerate(gstins, 1):
            self.update_status(f"Processing {gstin} ({i}/{len(gstins)})...")
            row = self.fetch_single(gstin)
            if row:
                all_rows.append(row)

        if not all_rows:
            self.update_status("No data extracted.")
            return

        fp = self.save_excel(all_rows)
        self.update_status("Done.")
        messagebox.showinfo("Saved", f"Data saved to:\n{fp}")

    # ===================== SCRAPER MAIN FUNCTION =====================

     def fetch_single(self, gstin):
        try:
            self.driver.get(GST_SEARCH_URL)
        except:
            messagebox.showerror("Error", "Cannot load GST portal.")
            return None

        time.sleep(2)

        # Enter GSTIN
        try:
            by, val = LOCATORS["gstin_input"]
            ele = self.driver.find_element(by, val)
            ele.clear()
            ele.send_keys(gstin)
        except:
            messagebox.showerror("Error", "GSTIN input not found.")
            return None

        # User manually enters CAPTCHA and clicks first Search
        messagebox.showinfo(
            "CAPTCHA",
            f"GSTIN: {gstin}\n\n"
            "Enter CAPTCHA manually and click the FIRST Search button.\n"
            "Wait until taxpayer details page loads.\n\n"
            "Then click OK."
        )

        time.sleep(3)

        # Scroll down near filing section
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight * 0.65);")
        time.sleep(1.5)

        # Auto click "Show Filing Table"
        try:
            btn = self.driver.find_element(*LOCATORS["show_filing_table_btn"])
            self.driver.execute_script("arguments[0].click();", btn)
            time.sleep(2)
        except:
            pass

        # Auto click filing section "Search"
        try:
            btn2 = self.driver.find_element(*LOCATORS["filing_search_btn"])
            self.driver.execute_script("arguments[0].click();", btn2)
            time.sleep(3)
        except:
            pass

        # Scroll fully to make tables visible
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(2)

        # Prepare return dict
        data = {
            "GSTIN": gstin,
            "Legal Name": safe_get_text(self.driver, *LOCATORS["legal_name"]),
            "Trade Name": safe_get_text(self.driver, *LOCATORS["trade_name"]),
            "Return Filing Status": "Latest filing dates shown",
            "GSTR-1 Info": "",
            "GSTR-3B Info": "",
            "Registration Status": safe_get_text(self.driver, *LOCATORS["reg_status"]),
        }

        # Latest GSTR-3B
        try:
            tbl = self.driver.find_element(*LOCATORS["gstr3b_info"])
            first = tbl.find_element(By.CSS_SELECTOR, "tbody tr")
            cols = [td.text.strip() for td in first.find_elements(By.TAG_NAME, "td")]
            data["GSTR-3B Info"] = cols[2] if len(cols) > 2 else ""
        except:
            data["GSTR-3B Info"] = ""

        # Latest GSTR-1
        try:
            tbl = self.driver.find_element(*LOCATORS["gstr1_info"])
            first = tbl.find_element(By.CSS_SELECTOR, "tbody tr")
            cols = [td.text.strip() for td in first.find_elements(By.TAG_NAME, "td")]
            data["GSTR-1 Info"] = cols[2] if len(cols) > 2 else ""
        except:
            data["GSTR-1 Info"] = ""

        return data

     def on_exit(self):
        if self.driver:
            try:
                self.driver.quit()
            except:
                pass
        self.root.destroy()


# ===================== RUN =====================

if __name__ == "__main__":
    root = tk.Tk()
    GSTFilingApp(root)
    root.mainloop()
